#!/usr/bin/python
##########################################################################################################
#        耦合系统 by：GCX
##########################################################################################################
import os
import numpy as np
import math
import h5py
import sys
import re
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import shutil
##########################################################################################################
#        耦合程序输入卡
##########################################################################################################
relaxation_factor = 1.0      #picard迭代松弛因子 1:直接代值，0.5：中值
fuel_temp_convergence_limit = 5
coolant_temp_convergence_limit = 1
power_convergence_limit = 0.005
#####################################################
temp_c = 0.0 #COBRA_pb为273.15，en为0.0
fuel_T0 = 900.0 #初始化燃料温度值
coolant_T0 = 565.0  #初始化冷却剂温度值
coolant_D0 = 0.743  #初始化冷却剂密度，单位g/cm3
liner_power_start_insert_point = "$ card 5" #填入记得空格  $+空格+card+number
liner_power_end_insert_point = "$  card 7" #填入记得两个空格
core_power = 3411000000/4 #总功率
#n_depl = 264*56*9 #燃耗区数目
n_pin_in_assembly = 264
n_eq_pin_in_assembly = 1 #组件模式采用集中参数为1，棒模式为组件内实际棒的个数
n_edge_channels_in_assembly = 0 #铅冷快堆组件内边通道数，压水堆对应GT外的冷却剂通道
n_corner_channels_in_assembly = 0 #铅冷快堆组件内角通道数，压水堆对应IT外的冷却剂通道
n_fuel_assembly = 56 #单层组件数目
#由于cobra必须输入完整的棒功率，但涉及如四分之一对称模型有半棒，或半组件，openmc计算的四分之一模型功率仅输出一半的，所以对应棒位置要乘2,但是组件无此操作
#quarter_assembly_indices = [1]
# quarter_assembly_indices = []
# quarter_assembly_indices = [x - 1 for x in quarter_assembly_indices]
#half_assembly_indices = [2,3,4,5,6,7,8,9,17,25,33,40,47,53]
# half_assembly_indices = []
# half_assembly_indices = [x - 1 for x in half_assembly_indices]
n_fuel_diff_type= [20,19,17] #单层燃料组件123的个数
#n_nofuel_assembly = 0
n_floor = 10
n_type_fuel = 3
#core_map_1 = [[3]*4,[3]*6,[3]*3,[2]*4,[3]*3,[3]*2,[2]*7,[3]*2,[3]*1,[2]*6,[3]*1,[3]*2,[2]*3,[1]*3,[2]*3,[3]*2,[3]*1,[2]*3,[1]*4,[2]*3,[3]*1,[3]*2,[2]*1,[1]*4,[2]*1,[3]*2,[3]*1,[2]*3,[1]*4,[2]*3,[3]*1,[3]*2,[2]*3,[1]*3,[2]*3,[3]*2,[3]*1,[2]*6,[3]*1,[3]*2,[2]*7,[3]*2,[3]*3,[2]*4,[3]*3,[3]*6,[3]*4]
#core_map = [int(x) for item in core_map_1 for x in item]
core_map = [3,3,3,3,
1,3,1,3,3,3,
2,1,2,1,2,3,3,
1,2,1,2,2,2,3,
2,1,2,1,2,1,3,3,
1,2,1,2,1,2,1,3,
2,1,2,1,2,1,3,3,
1,2,1,2,1,2,1,3
]

#core_map = [3,3,3,]#nofuel就填0，范围0-5，同时最大值小于n_type_fuel
fuel_1_start_tally_number = 1 #第一层的第一种燃料棒级tally开始的号，对应tally8为第一层第二种，以此类推，层层向上
'''
z_pin = [0.01933,0.07972,0.16183,0.24394,0.32605,0.40816,0.49027,0.57238,
0.63249,0.69186,0.77251,0.85316,0.93382,1.01446,1.09512,1.15449,1.21386,
1.29452,1.37516,1.45582,1.53646,1.61712,1.67649,1.73586,1.81651,1.89716,
1.97781,2.05846,2.13911,2.19849,2.25786,2.33851,2.41917,2.49981,2.58046,
2.66112,2.72049,2.77987,2.86052,2.94117,3.02182,3.10247,3.18312,3.24249,
3.30115,3.38036,3.45957,3.53878,3.59819,3.63800]   #每个节块中心的z轴坐标
# h_pin_1 = [[0.18288/1]*20]
# h_pin = [float(x) for item in h_pin_1 for x in item]   #目前使用相同高度节块
h_pin = [3.86600e-02,8.21110e-02,8.21120e-02,8.21110e-02,8.21120e-02,8.21110e-02,
8.21120e-02,8.21110e-02,3.81000e-02,8.06500e-02,8.06500e-02,8.06500e-02,
8.06500e-02,8.06500e-02,8.06500e-02,3.81000e-02,8.06500e-02,8.06500e-02,
8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,3.81000e-02,8.06500e-02,
8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,3.81000e-02,
8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,
3.81000e-02,8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,8.06500e-02,
8.06500e-02,3.81000e-02,7.92120e-02,7.92120e-02,7.92120e-02,7.92120e-02,
3.96060e-02,3.96060e-02]
'''

# 修改后的节块高度 (10个相同的0.36576m)
h_pin = [
    3.65760e-01, 3.65760e-01, 3.65760e-01, 3.65760e-01, 3.65760e-01,
    3.65760e-01, 3.65760e-01, 3.65760e-01, 3.65760e-01, 3.65760e-01
]

# 修改后的节块中心z轴坐标 (0~3.65760m等分10层的中心)
z_pin = [
    0.18288,  # 第1层中心 = (0 + 0.36576/2)
    0.54864,  # 第2层中心 = (0.36576 + 0.36576/2)
    0.91440,  # 第3层中心
    1.28016,  # 第4层中心
    1.64592,  # 第5层中心
    2.01168,  # 第6层中心
    2.37744,  # 第7层中心
    2.74320,  # 第8层中心
    3.10896,  # 第9层中心
    3.47472   # 第10层中心 = (3.65760 - 0.36576/2)
]
###########################################################
run_openmc ="mpirun --bind-to none -host node2:1,node3:1,node4:1,node5:1,node6:1,node7:1,node8:1,node12:1,node13:1,node14:1,node15:1,node16:1 openmc "#> /dev/null 2>&1
run_cobra = "./EN"
coupled_mode = "assembly"
cobra_input_file = "INPFILE"
cobra_out_file = "OUTFILE"
fuel_file = 'OUTFILE'
coolant_file ='OUTFILE'
statepoint_file = "statepoint.300.h5" #点状态文件名
fuel_1_name = 'fuel_1'
fuel_2_name = 'fuel_2'
fuel_3_name = 'fuel_3'
fuel_4_name = 'fuel_4'
fuel_5_name = 'fuel_5'

fuel_1_assembly_inner_coolant_channals_name  =   'fuel_1_assembly_coolant_i'
fuel_2_assembly_inner_coolant_channals_name  =   'fuel_2_assembly_coolant_i'
fuel_3_assembly_inner_coolant_channals_name  =   'fuel_3_assembly_coolant_i'
fuel_4_assembly_inner_coolant_channals_name  =   'fuel_4_assembly_coolant_i'
fuel_5_assembly_inner_coolant_channals_name  =   'fuel_5_assembly_coolant_i'
                                                  
fuel_1_assembly_edge_coolant_channals_name   =   'fuel_1_assembly_coolant_e'
fuel_2_assembly_edge_coolant_channals_name   =   'fuel_2_assembly_coolant_e'
fuel_3_assembly_edge_coolant_channals_name   =   'fuel_3_assembly_coolant_e'
fuel_4_assembly_edge_coolant_channals_name   =   'fuel_4_assembly_coolant_e'
fuel_5_assembly_edge_coolant_channals_name   =   'fuel_5_assembly_coolant_e'
                                                  
fuel_1_assembly_corner_coolant_channals_name =   'fuel_1_assembly_coolant_c'
fuel_2_assembly_corner_coolant_channals_name =   'fuel_2_assembly_coolant_c'
fuel_3_assembly_corner_coolant_channals_name =   'fuel_3_assembly_coolant_c'
fuel_4_assembly_corner_coolant_channals_name =   'fuel_4_assembly_coolant_c'
fuel_5_assembly_corner_coolant_channals_name =   'fuel_5_assembly_coolant_c'
################################################################################################################
#      耦合程序图标
#################################################################################################################
print(r"""
						   %%%%%%%%%%%%%%%%%%%%%%%%
						%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
					  %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
					%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
				   %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
				  COBRA-EN ——————-\ %%%%%%%%%%%%%%%%%%%%%%%%
				 COBRA-PB -——————\ \ %%%%%%%%%%%%%%%%%%%%%%%%
				 ###############  \ \ %%%%%%%%%%%%%%%%%%%%%%%%
				################## \ \ %%%%%%%%%%%%%%%%%%%%%%%
				################### \ \ %%%%%%%%%%%%%%%%%%%%%%%
				#################### \ \ %%%%%%%%%%%%%%%%%%%%%%
				##################### \ \ %%%%%%%%%%%%%%%%%%%%%
				###################### \ \ %%%%%%%%%%%%%%%%%%%%
				####################### C C %%%%%%%%%%%%%%%%%%
				 ####################### O O %%%%%%%%%%%%%%%%%
				 ###################### B B %%%%%%%%%%%%%%%%%
				  #################### R R %%%%%%%%%%%%%%%%%
					################# A A %%%%%%%%%%%%%%%%%
					 ############### | | %%%%%%%%%%%%%%%%
					   ############ P E %%%%%%%%%%%%%%%
						  ######## B N %%%%%%%%%%%%%%
									  %%%%%%%%%%%

				 | The OpenMC And COBRA-EN/PB
	   Copyright | China Three Gorges University
		 Version | OpenMC-0.14.0 , COBRA-EN/PB
""")
#去燃耗的冷却剂通道名
no_depletion_coolant_names = [
fuel_1_assembly_inner_coolant_channals_name ,
fuel_2_assembly_inner_coolant_channals_name ,
fuel_3_assembly_inner_coolant_channals_name ,
fuel_4_assembly_inner_coolant_channals_name ,
fuel_5_assembly_inner_coolant_channals_name ,
fuel_1_assembly_edge_coolant_channals_name  ,
fuel_2_assembly_edge_coolant_channals_name  ,
fuel_3_assembly_edge_coolant_channals_name  ,
fuel_4_assembly_edge_coolant_channals_name  ,
fuel_5_assembly_edge_coolant_channals_name  ,
fuel_1_assembly_corner_coolant_channals_name,
fuel_2_assembly_corner_coolant_channals_name,
fuel_3_assembly_corner_coolant_channals_name,
fuel_4_assembly_corner_coolant_channals_name,
fuel_5_assembly_corner_coolant_channals_name,]
##############################################################################################################
#        燃料温度初始化
##################################################################################################################
if n_type_fuel == 1 :
	fuel_1_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_T0 = np.array([])
	fuel_3_assembly_T0 = np.array([])
	fuel_4_assembly_T0 = np.array([])
	fuel_5_assembly_T0 = np.array([])
elif n_type_fuel == 2 :
	fuel_1_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_T0 = np.array([])
	fuel_4_assembly_T0 = np.array([])
	fuel_5_assembly_T0 = np.array([])
elif n_type_fuel == 3 :
	fuel_1_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_T0 = np.array([])
	fuel_5_assembly_T0 = np.array([])
elif n_type_fuel == 4 :
	fuel_1_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[3] * n_floor*n_eq_pin_in_assembly)
	fuel_5_assembly_T0 = np.array([])
elif n_type_fuel == 5 :
	fuel_1_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[3] * n_floor*n_eq_pin_in_assembly)
	fuel_5_assembly_T0 = np.array([fuel_T0] * n_fuel_diff_type[4] * n_floor*n_eq_pin_in_assembly)
else :
	print('初始化燃料温度错误')
##############################################################################################################
#        冷却剂温度初始化
##################################################################################################################
if n_type_fuel == 1 :
	fuel_1_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_T0 = np.array([])
	fuel_3_assembly_coolant_T0 = np.array([])
	fuel_4_assembly_coolant_T0 = np.array([])
	fuel_5_assembly_coolant_T0 = np.array([])
elif n_type_fuel == 2 :
	fuel_1_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_T0 = np.array([])
	fuel_4_assembly_coolant_T0 = np.array([])
	fuel_5_assembly_coolant_T0 = np.array([])
elif n_type_fuel == 3 :
	fuel_1_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_coolant_T0 = np.array([])
	fuel_5_assembly_coolant_T0 = np.array([])
elif n_type_fuel == 4 :
	fuel_1_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[3] * n_floor*n_eq_pin_in_assembly)
	fuel_5_assembly_coolant_T0 = np.array([])
elif n_type_fuel == 5 :
	fuel_1_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[3] * n_floor*n_eq_pin_in_assembly)
	fuel_5_assembly_coolant_T0 = np.array([coolant_T0] * n_fuel_diff_type[4] * n_floor*n_eq_pin_in_assembly)
else :
	print('初始化冷却剂温度错误')
##############################################################################################################
#        冷却剂密度初始化
##################################################################################################################
if n_type_fuel == 1 :
	fuel_1_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_D0 = np.array([])
	fuel_3_assembly_coolant_D0 = np.array([])
	fuel_4_assembly_coolant_D0 = np.array([])
	fuel_5_assembly_coolant_D0 = np.array([])
	if coupled_mode == "pin":
		fuel_1_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_2_assembly_coolant_e_D0 = np.array([])
		fuel_3_assembly_coolant_e_D0 = np.array([])
		fuel_4_assembly_coolant_e_D0 = np.array([])
		fuel_5_assembly_coolant_e_D0 = np.array([])
		fuel_1_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_2_assembly_coolant_c_D0 = np.array([])
		fuel_3_assembly_coolant_c_D0 = np.array([])
		fuel_4_assembly_coolant_c_D0 = np.array([])
		fuel_5_assembly_coolant_c_D0 = np.array([])
elif n_type_fuel == 2 :
	fuel_1_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_D0 = np.array([])
	fuel_4_assembly_coolant_D0 = np.array([])
	fuel_5_assembly_coolant_D0 = np.array([])
	if coupled_mode == "pin":
		fuel_1_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_2_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_3_assembly_coolant_e_D0 = np.array([])
		fuel_4_assembly_coolant_e_D0 = np.array([])
		fuel_5_assembly_coolant_e_D0 = np.array([])
		fuel_1_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_2_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_3_assembly_coolant_c_D0 = np.array([])
		fuel_4_assembly_coolant_c_D0 = np.array([])
		fuel_5_assembly_coolant_c_D0 = np.array([])
elif n_type_fuel == 3 :
	fuel_1_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_coolant_D0 = np.array([])
	fuel_5_assembly_coolant_D0 = np.array([])
	if coupled_mode == "pin":
		fuel_1_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_2_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_3_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_4_assembly_coolant_e_D0 = np.array([])
		fuel_5_assembly_coolant_e_D0 = np.array([])
		fuel_1_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_2_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_3_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_4_assembly_coolant_c_D0 = np.array([])
		fuel_5_assembly_coolant_c_D0 = np.array([])
elif n_type_fuel == 4 :
	fuel_1_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[3] * n_floor*n_eq_pin_in_assembly)
	fuel_5_assembly_coolant_D0 = np.array([])
	if coupled_mode == "pin":
		fuel_1_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_2_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_3_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_4_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_5_assembly_coolant_e_D0 = np.array([])
		fuel_1_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_2_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_3_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_4_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_5_assembly_coolant_c_D0 = np.array([])
elif n_type_fuel == 5 :
	fuel_1_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[0] * n_floor*n_eq_pin_in_assembly)
	fuel_2_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[1] * n_floor*n_eq_pin_in_assembly)
	fuel_3_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[2] * n_floor*n_eq_pin_in_assembly)
	fuel_4_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[3] * n_floor*n_eq_pin_in_assembly)
	fuel_5_assembly_coolant_D0 = np.array([coolant_D0] * n_fuel_diff_type[4] * n_floor*n_eq_pin_in_assembly)
	if coupled_mode == "pin":
		fuel_1_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_2_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_3_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_4_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_5_assembly_coolant_e_D0 = np.array([coolant_D0] * n_floor * n_edge_channels_in_assembly)
		fuel_1_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_2_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_3_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_4_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
		fuel_5_assembly_coolant_c_D0 = np.array([coolant_D0] * n_floor * n_corner_channels_in_assembly)
else :
	print('初始化冷却剂密度错误')
##################################################################################################################
#        功率初始化使用平均功率
##################################################################################################################
assembly_powers0 = np.full((int(n_fuel_assembly), n_floor), core_power/n_floor/n_fuel_assembly)
##################################################################################################################
#      openmc初步运行试算
##################################################################################################################

#材料卡中的冷却剂去燃耗属性
print('======================材料数据初始化中=======================')
tree = ET.parse('materials.xml')
root = tree.getroot()
for material in root.findall('material'):
	if material.get('name') in no_depletion_coolant_names:
		# 删除 depletable 属性（如果存在）
		if 'depletable' in material.attrib:
			del material.attrib['depletable']

		# 删除 volume 属性（如果存在）
		if 'volume' in material.attrib:
			del material.attrib['volume']
#写入到新文件
tree.write('materials.xml', encoding='utf-8')

print('======================物理数据初始化中=======================')
os.system(run_openmc)
##################################################################################################################
#建立一维功率系数数组
##################################################################################################################
with h5py.File(statepoint_file, 'r') as hdf5_file:
	#data = np.zeros((max(n_fuel_diff_type) * n_eq_pin_in_assembly, n_floor * n_type_fuel))
	data_0 = np.array(hdf5_file['tallies/tally '+str(fuel_1_start_tally_number)+'/results'][:,0])[:,0]
	data = data_0[data_0 != 0]#去除0元素

	# rows = data.shape
	# print(data)
	# print(f"数组有 {rows} (行，列)")
##########################################################
#一维列功率系数数组归一化为一维功率数组
##########################################################
	data_c = data   #rows行，columns为列
	data_c = np.array(data_c)

	data_c /= np.mean(data_c[data_c > 0.]) #归一化，note：燃耗区不用相等，由于openmc计数卡计数的是不同区域的裂变次数，且仅需归一化出不同区域的功率
	data_c = core_power/(n_floor*n_fuel_assembly)*data_c #节块功率

	# data_c = np.array([f"{x:.5e}" for x in data_c])	#格式化数据
	# rows = data_c.shape
	# print(data_c)
	# print(f"数组有 {rows} 行，列")
##################################################################################################
#组件线功率分层写入cobra输入卡
##################################################################################################
floor = 1
i = 0
assembly_powers = np.zeros((int(len(data_c)/n_floor), n_floor))
assembly_liner_powers = np.zeros((int(len(data_c)/n_floor), n_floor))
while floor <= n_floor :
	assembly_powers[:,floor-1] = data_c[int(i):int(i+len(data_c)/n_floor)]
	assembly_liner_powers[:,floor-1] = assembly_powers[:,floor-1]/h_pin[floor-1]
	#assembly_liner_powers[:,floor-1][half_assembly_indices] *= 2
	#assembly_liner_powers[:,floor-1][quarter_assembly_indices] *= 4
	i += len(data_c)/n_floor
	floor += 1

assembly_liner_powers = np.array([[f"{x:.5e}" for x in row] for row in assembly_liner_powers])  # 注意：此为二维数据格式化,科学计数法且保留五位小数

#print(assembly_liner_powers[:,0])

print('耦合模式：'+str(coupled_mode))


with open(cobra_input_file, "r") as file:
	lines = file.readlines()


########################################################################  assembly
# 找到 "Start Line" 和 "End Line" 的位置
start_index = None
end_index = None
for i, line in enumerate(lines):
	if line.strip() == liner_power_start_insert_point:
		start_index = i
	elif line.strip() == liner_power_end_insert_point:
		end_index = i
if start_index is None or end_index is None:  # 检查是否找到两行
	print("未找到 'Start Line' 或 'End Line'")
else:
	floor = 1
	new_content = ""  # 构造新的内容
	new_content = new_content + "    "+str(n_floor)+ "\n"
	while floor <= n_floor:
		new_content = new_content + "  " + str(z_pin[floor-1]) + "\n"
		for i, item in enumerate(assembly_liner_powers[:, floor - 1]):
			new_content += str(item) + " "
			if (i + 1) % 6 == 0:  # 每六个元素一换行
				new_content += "\n"
		if n_fuel_assembly % 6 != 0:
			new_content += "\n"
		floor += 1
	lines[start_index + 1:end_index] = new_content

with open(cobra_input_file, "w") as file:  # 写入输出文件
	file.writelines(lines)
	print(f"功率信息载入...")

#####################################################################################################
#              运行cobra
#####################################################################################################
print('=====================热工数据初始化中========================')
os.system(run_cobra)
#####################################################################################################
#              读取燃料温度建立数组
#####################################################################################################

# 读取文件内容
with open(fuel_file, 'r') as file:
	file_content = file.read()

# 初始化变量
rods_data = {}  # 存储每个燃料棒的温度数据
current_rod = None
height_positions = []  # 存储所有高度位置

# 正则表达式模式
rod_pattern = re.compile(r'TEMPERATURE DATA FOR ROD\s+(\d+) \(FUEL TYPE\s+\d+\)')
temp_pattern = re.compile(r'\s*(\d+\.\d+)\s+\d+\.\d+\s+\d+\.\d+\s+\d+\s+(\d+\.\d+)')

# 处理每一行
for line in file_content.split('\n'):
# 检查是否是新的燃料棒数据开始
	rod_match = rod_pattern.search(line)
	if rod_match:
		current_rod = int(rod_match.group(1))
		rods_data[current_rod] = {}  # 初始化该燃料棒的数据字典
		continue

# 检查是否是温度数据行
	temp_match = temp_pattern.match(line)
	if temp_match and current_rod is not None:
		height = temp_match.group(1)
		temperature = float(temp_match.group(2))

		# 记录高度位置（如果尚未记录）
		if height not in height_positions:
			height_positions.append(height)

		rods_data[current_rod][height] = temperature  # 存储该燃料棒在该高度的温度

# 按高度排序（从小到大）
height_positions = sorted(height_positions, key=float)

# 按燃料棒编号排序（从小到大）
rod_numbers = sorted(rods_data.keys())

# 构建二维数组（行=燃料棒，列=高度）
temperature_2d_array = []
for rod in rod_numbers:
	rod_temps = []
	for height in height_positions:
		rod_temps.append(rods_data[rod][height])
	temperature_2d_array.append(rod_temps)

#温度单位由摄氏度转为K
fuel_temperature_2d_array = np.array(temperature_2d_array)
fuel_temperature_2d_array = np.array([temp + temp_c for temp in fuel_temperature_2d_array])
#print(fuel_temperature_2d_array)

#####################################################################################################
#      读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理
#####################################################################################################
floor = 1
fuel_1_assembly_T = []   #rows行，columns为列
fuel_2_assembly_T = []   #rows行，columns为列
fuel_3_assembly_T = []   #rows行，columns为列
fuel_4_assembly_T = []   #rows行，columns为列
fuel_5_assembly_T = []   #rows行，columns为列
while floor <= n_floor :
	i = 0
	j = 0
	while i <= len(core_map)-1 :
		if core_map[i] == 0:
			print("have no fuel region")
		elif core_map[i] == 1:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = fuel_temperature_2d_array[j, floor - 1]
				fuel_1_assembly_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 2:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = fuel_temperature_2d_array[j,floor-1]
				fuel_2_assembly_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 3:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = fuel_temperature_2d_array[j,floor-1]
				fuel_3_assembly_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 4:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = fuel_temperature_2d_array[j,floor-1]
				fuel_4_assembly_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 5:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = fuel_temperature_2d_array[j,floor-1]
				fuel_5_assembly_T.append(subset)
				k += 1
				j += 1
		else:
			print(f"读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理错误")
			sys.exit(1)
		i += 1
	floor += 1
fuel_1_assembly_T = np.array(fuel_1_assembly_T)
fuel_2_assembly_T = np.array(fuel_2_assembly_T)
fuel_3_assembly_T = np.array(fuel_3_assembly_T)
fuel_4_assembly_T = np.array(fuel_4_assembly_T)
fuel_5_assembly_T = np.array(fuel_5_assembly_T)

#####################################################################################################
#              读取冷却剂温度和密度建立数组
#####################################################################################################
# 读取文件内容
with open(coolant_file, 'r') as file:
	content = file.read()

# 初始化数据结构
channels = {}

# 使用正则表达式分割不同通道的数据
channel_sections = re.split(r'TIME =\s+\d+\.\d+ SEC\s+-\s+RESULTS FOR CHANNEL\s+\d+', content)[1:]

# 遍历每个通道的数据
for i, section in enumerate(channel_sections, start=1):
	# 初始化当前通道的数据列表
	channels[i] = {'distance': [], 'temperature': [], 'density': []}

	# 分割行并跳过标题行
	lines = [line.strip() for line in section.split('\n') if line.strip()]
	data_lines = lines[2:]  # 跳过前两行（标题和空行）

	# 提取每行的数据
	for line in data_lines:
		# 使用正则表达式匹配数据行
		match = re.match(r'\s*(\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)', line)
		if match:
			distance = float(match.group(1))
			temperature = float(match.group(4))
			density = float(match.group(5))

			channels[i]['distance'].append(distance)
			channels[i]['temperature'].append(temperature)
			channels[i]['density'].append(density)

# 创建二维数组（假设所有通道有相同数量的数据点）
num_channels = len(channels)
num_points = len(channels[1]['temperature'])

# 温度二维数组 (channels × points)
coolant_temperature_2d = []
for ch in range(1, num_channels + 1):
	coolant_temperature_2d.append([channels[ch]['temperature'][i] for i in range(num_points)])

# 密度二维数组 (channels × points)
coolant_density_2d = []
for ch in range(1, num_channels + 1):
	coolant_density_2d.append([channels[ch]['density'][i] for i in range(num_points)])


#转nump数组温度单位由摄氏度转为K
coolant_temperature_2d_array = np.array(coolant_temperature_2d)
coolant_temperature_2d_array = np.array([temp + temp_c for temp in coolant_temperature_2d_array])
#转nump数组
coolant_density_2d_array = np.array(coolant_density_2d)

#平均值法求解节块中心温度
coolant_temperature_2d_array = (coolant_temperature_2d_array[:, :-1] + coolant_temperature_2d_array[:, 1:]) / 2
#平均值法求解节块中心密度
coolant_density_2d_array = (coolant_density_2d_array[:, :-1] + coolant_density_2d_array[:, 1:]) / 2
#密度单位kg/m3改为g/cm3
coolant_density_2d_array = coolant_density_2d_array /1000
#####################################################################################################
#      读取冷却剂温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理
#####################################################################################################
floor = 1
fuel_1_assembly_coolant_T = []   #rows行，columns为列
fuel_2_assembly_coolant_T = []   #rows行，columns为列
fuel_3_assembly_coolant_T = []   #rows行，columns为列
fuel_4_assembly_coolant_T = []   #rows行，columns为列
fuel_5_assembly_coolant_T = []   #rows行，columns为列
while floor <= n_floor :
	i = 0
	j = 0
	while i <= len(core_map)-1 :
		if core_map[i] == 0:
			print("have no fuel region")
		elif core_map[i] == 1:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_temperature_2d_array[j,floor-1]
				fuel_1_assembly_coolant_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 2:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_temperature_2d_array[j,floor-1]
				fuel_2_assembly_coolant_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 3:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_temperature_2d_array[j,floor-1]
				fuel_3_assembly_coolant_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 4:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_temperature_2d_array[j,floor-1]
				fuel_4_assembly_coolant_T.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 5:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_temperature_2d_array[j,floor-1]
				fuel_5_assembly_coolant_T.append(subset)
				k += 1
				j += 1
		else:
			print(f"读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理错误")
			sys.exit(1)
		i += 1
	floor += 1
fuel_1_assembly_coolant_T = np.array(fuel_1_assembly_coolant_T)
fuel_2_assembly_coolant_T = np.array(fuel_2_assembly_coolant_T)
fuel_3_assembly_coolant_T = np.array(fuel_3_assembly_coolant_T)
fuel_4_assembly_coolant_T = np.array(fuel_4_assembly_coolant_T)
fuel_5_assembly_coolant_T = np.array(fuel_5_assembly_coolant_T)

#####################################################################################################
#      读取冷却剂密度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理
#####################################################################################################
floor = 1
fuel_1_assembly_coolant_D = []   #rows行，columns为列
fuel_2_assembly_coolant_D = []   #rows行，columns为列
fuel_3_assembly_coolant_D = []   #rows行，columns为列
fuel_4_assembly_coolant_D = []   #rows行，columns为列
fuel_5_assembly_coolant_D = []   #rows行，columns为列
while floor <= n_floor :
	i = 0
	j = 0
	while i <= len(core_map)-1 :
		if core_map[i] == 0:
			print("have no fuel region")
		elif core_map[i] == 1:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_density_2d_array[j,floor-1]
				fuel_1_assembly_coolant_D.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 2:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_density_2d_array[j,floor-1]
				fuel_2_assembly_coolant_D.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 3:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_density_2d_array[j,floor-1]
				fuel_3_assembly_coolant_D.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 4:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_density_2d_array[j,floor-1]
				fuel_4_assembly_coolant_D.append(subset)
				k += 1
				j += 1
		elif core_map[i] == 5:
			k = 0
			while k < n_eq_pin_in_assembly:
				subset = coolant_density_2d_array[j,floor-1]
				fuel_5_assembly_coolant_D.append(subset)
				k += 1
				j += 1
		else:
			print(f"读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理错误")
			sys.exit(1)
		i += 1
	floor += 1
fuel_1_assembly_coolant_D = np.array(fuel_1_assembly_coolant_D)
fuel_2_assembly_coolant_D = np.array(fuel_2_assembly_coolant_D)
fuel_3_assembly_coolant_D = np.array(fuel_3_assembly_coolant_D)
fuel_4_assembly_coolant_D = np.array(fuel_4_assembly_coolant_D)
fuel_5_assembly_coolant_D = np.array(fuel_5_assembly_coolant_D)
#print(fuel_1_assembly_coolant_D)
#####################################################################################################
#      燃料和冷却剂信息转存
#####################################################################################################
fuel_1_assembly_T1 = fuel_1_assembly_T
fuel_2_assembly_T1 = fuel_2_assembly_T
fuel_3_assembly_T1 = fuel_3_assembly_T
fuel_4_assembly_T1 = fuel_4_assembly_T
fuel_5_assembly_T1 = fuel_5_assembly_T

fuel_1_assembly_coolant_T1 = fuel_1_assembly_coolant_T
fuel_2_assembly_coolant_T1 = fuel_2_assembly_coolant_T
fuel_3_assembly_coolant_T1 = fuel_3_assembly_coolant_T
fuel_4_assembly_coolant_T1 = fuel_4_assembly_coolant_T
fuel_5_assembly_coolant_T1 = fuel_5_assembly_coolant_T

fuel_1_assembly_coolant_D1 = fuel_1_assembly_coolant_D
fuel_2_assembly_coolant_D1 = fuel_2_assembly_coolant_D
fuel_3_assembly_coolant_D1 = fuel_3_assembly_coolant_D
fuel_4_assembly_coolant_D1 = fuel_4_assembly_coolant_D
fuel_5_assembly_coolant_D1 = fuel_5_assembly_coolant_D
#print(fuel_1_assembly_T)
'''
#####################################################################################################
#      燃料和冷却剂的picard迭代
#####################################################################################################
fuel_1_assembly_T = fuel_1_assembly_T0 + relaxation_factor*(fuel_1_assembly_T1 - fuel_1_assembly_T0)
fuel_2_assembly_T = fuel_2_assembly_T0 + relaxation_factor*(fuel_2_assembly_T1 - fuel_2_assembly_T0)
fuel_3_assembly_T = fuel_3_assembly_T0 + relaxation_factor*(fuel_3_assembly_T1 - fuel_3_assembly_T0)
fuel_4_assembly_T = fuel_4_assembly_T0 + relaxation_factor*(fuel_4_assembly_T1 - fuel_4_assembly_T0)
fuel_5_assembly_T = fuel_5_assembly_T0 + relaxation_factor*(fuel_5_assembly_T1 - fuel_5_assembly_T0)

fuel_1_assembly_coolant_T = fuel_1_assembly_coolant_T0 + relaxation_factor*(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0)
fuel_2_assembly_coolant_T = fuel_2_assembly_coolant_T0 + relaxation_factor*(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0)
fuel_3_assembly_coolant_T = fuel_3_assembly_coolant_T0 + relaxation_factor*(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0)
fuel_4_assembly_coolant_T = fuel_4_assembly_coolant_T0 + relaxation_factor*(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0)
fuel_5_assembly_coolant_T = fuel_5_assembly_coolant_T0 + relaxation_factor*(fuel_5_assembly_coolant_T1 - fuel_5_assembly_coolant_T0)

fuel_1_assembly_coolant_D = fuel_1_assembly_coolant_D0 + relaxation_factor*(fuel_1_assembly_coolant_D1 - fuel_1_assembly_coolant_D0)
fuel_2_assembly_coolant_D = fuel_2_assembly_coolant_D0 + relaxation_factor*(fuel_2_assembly_coolant_D1 - fuel_2_assembly_coolant_D0)
fuel_3_assembly_coolant_D = fuel_3_assembly_coolant_D0 + relaxation_factor*(fuel_3_assembly_coolant_D1 - fuel_3_assembly_coolant_D0)
fuel_4_assembly_coolant_D = fuel_4_assembly_coolant_D0 + relaxation_factor*(fuel_4_assembly_coolant_D1 - fuel_4_assembly_coolant_D0)
fuel_5_assembly_coolant_D = fuel_5_assembly_coolant_D0 + relaxation_factor*(fuel_5_assembly_coolant_D1 - fuel_5_assembly_coolant_D0)
'''

print('======================数据初始化完成=========================')

#############################################################################################################################################################################################
#############################################################################################################################################################################################
#############################################################################################################################################################################################
iterations = 1
# 定义各个条件
if n_type_fuel == 1 :
	fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_temp_condition_max = fuel_condition1
	coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,
										ord=np.inf) > coolant_temp_convergence_limit
	coolant_temp_condition_max = coolant_condition1

	fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1,))
	fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0,))
	fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
	fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
	coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1,))
	coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0,))
	coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
	coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
	max_fuel_convergence = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf)
	max_coolant_convergence = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf)
elif n_type_fuel == 2 :
	fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_temp_condition_max = fuel_condition1 or fuel_condition2
	coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_temp_condition_max = coolant_condition1 or coolant_condition2

	fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1, fuel_2_assembly_T1))
	fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0, fuel_2_assembly_T0))
	fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
	fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
	coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1))
	coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0))
	coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
	coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
	max_fuel_convergence = max(
	np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf))
	max_coolant_convergence = max(
	np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf))
elif n_type_fuel == 3 :
	fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition3 = np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_temp_condition_max = fuel_condition1 or fuel_condition2 or fuel_condition3
	coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition3 = np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_temp_condition_max = coolant_condition1 or coolant_condition2 or coolant_condition3

	fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1, fuel_2_assembly_T1, fuel_3_assembly_T1))
	fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0, fuel_2_assembly_T0, fuel_3_assembly_T0))
	fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
	fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
	coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1, fuel_3_assembly_coolant_T1))
	coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0, fuel_3_assembly_coolant_T0))
	coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
	coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
	max_fuel_convergence = max(
	np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf))
	max_coolant_convergence = max(
	np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0, ord=np.inf))
elif n_type_fuel == 4 :
	fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition3 = np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition4 = np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_temp_condition_max = fuel_condition1 or fuel_condition2 or fuel_condition3 or fuel_condition4
	coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition3 = np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition4 = np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_temp_condition_max = coolant_condition1 or coolant_condition2 or coolant_condition3 or coolant_condition4

	fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1, fuel_2_assembly_T1, fuel_3_assembly_T1, fuel_4_assembly_T1))
	fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0, fuel_2_assembly_T0, fuel_3_assembly_T0, fuel_4_assembly_T0))
	fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
	fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
	coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1, fuel_3_assembly_coolant_T1, fuel_4_assembly_coolant_T1))
	coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0, fuel_3_assembly_coolant_T0, fuel_4_assembly_coolant_T0))
	coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
	coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
	max_fuel_convergence = max(
	np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0, ord=np.inf))
	max_coolant_convergence = max(
	np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0, ord=np.inf))
elif n_type_fuel == 5 :
	fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition3 = np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition4 = np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_condition5 = np.linalg.norm(fuel_5_assembly_T1 - fuel_5_assembly_T0, ord=np.inf) > fuel_temp_convergence_limit
	fuel_temp_condition_max = fuel_condition1 or fuel_condition2 or fuel_condition3 or fuel_condition4 or fuel_condition5
	coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition3 = np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition4 = np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_condition5 = np.linalg.norm(fuel_5_assembly_coolant_T1 - fuel_5_assembly_coolant_T0,ord=np.inf) > coolant_temp_convergence_limit
	coolant_temp_condition_max = coolant_condition1 or coolant_condition2 or coolant_condition3 or coolant_condition4 or coolant_condition5

	fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1, fuel_2_assembly_T1, fuel_3_assembly_T1, fuel_4_assembly_T1, fuel_5_assembly_T1))
	fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0, fuel_2_assembly_T0, fuel_3_assembly_T0, fuel_4_assembly_T0, fuel_5_assembly_T0))
	fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
	fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
	coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1,fuel_3_assembly_coolant_T1, fuel_4_assembly_coolant_T1, fuel_5_assembly_coolant_T1))
	coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0,fuel_3_assembly_coolant_T0, fuel_4_assembly_coolant_T0, fuel_5_assembly_coolant_T0))
	coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
	coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
	max_fuel_convergence = max(
	np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0, ord=np.inf),
	np.linalg.norm(fuel_5_assembly_T1 - fuel_5_assembly_T0, ord=np.inf))
	max_coolant_convergence = max(
	np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0, ord=np.inf),
	np.linalg.norm(fuel_5_assembly_coolant_T1 - fuel_5_assembly_coolant_T0, ord=np.inf))
else :
	print("收敛条件处理错误")


abs_error = np.abs(assembly_powers - assembly_powers0)
eps = 1e-12
relative_error = abs_error / (np.abs(assembly_powers0) + eps)  # 计算相对误差（避免除以零）
power_rmse = np.sqrt(np.mean(relative_error ** 2))

power_condition_max = np.max(relative_error) > power_convergence_limit
power_condition = power_rmse > power_convergence_limit

print("最大范数数据处理：")
print(" 燃料温度收敛判据:", max_fuel_convergence)
print(" 冷却剂温度收敛判据:",max_coolant_convergence)
print(" 功率收敛判据:",np.max(relative_error))

print("均方根数据处理：")
print(" 燃料温度收敛判据:", fuel_temp_rmse)
print(" 冷却剂温度收敛判据:",coolant_temp_rmse)
print(" 功率收敛判据:",power_rmse)

directory = "./" + "迭代步"+str(0)+"输出"
if not os.path.exists(directory):  # 文件夹存在检测并创建
	os.makedirs(directory)
all_convergence = np.array([[max_fuel_convergence,fuel_temp_rmse],
					   [max_coolant_convergence,coolant_temp_rmse],
					   [np.max(relative_error),power_rmse]])
np.savetxt("./" + "迭代步" + str(0) + "输出" + "/收敛判据.csv", all_convergence, delimiter=', ')

all_fuel_T_xH_yN = []  #防报错

while  fuel_temp_condition_max or coolant_temp_condition_max or power_condition_max :
	print('\n======================第'+str(iterations)+'次迭代开始==========================')
	# 温度，功率和密度储存
	assembly_powers0 = assembly_powers

	fuel_1_assembly_T0 = fuel_1_assembly_T1
	fuel_2_assembly_T0 = fuel_2_assembly_T1
	fuel_3_assembly_T0 = fuel_3_assembly_T1
	fuel_4_assembly_T0 = fuel_4_assembly_T1
	fuel_5_assembly_T0 = fuel_5_assembly_T1

	fuel_1_assembly_coolant_T0 = fuel_1_assembly_coolant_T1
	fuel_2_assembly_coolant_T0 = fuel_2_assembly_coolant_T1
	fuel_3_assembly_coolant_T0 = fuel_3_assembly_coolant_T1
	fuel_4_assembly_coolant_T0 = fuel_4_assembly_coolant_T1
	fuel_5_assembly_coolant_T0 = fuel_5_assembly_coolant_T1

	fuel_1_assembly_coolant_D0 = fuel_1_assembly_coolant_D1
	fuel_2_assembly_coolant_D0 = fuel_2_assembly_coolant_D1
	fuel_3_assembly_coolant_D0 = fuel_3_assembly_coolant_D1
	fuel_4_assembly_coolant_D0 = fuel_4_assembly_coolant_D1
	fuel_5_assembly_coolant_D0 = fuel_5_assembly_coolant_D1


	# 前一次的热工计算的二维温度信息转存，也即模拟openmc使用的温度信息
	all_fuel_T_xH_yN = fuel_temperature_2d_array
	all_coolant_T_xH_yN = coolant_temperature_2d_array
	all_coolant_D_xH_yN = coolant_density_2d_array
	#####################################################################################################
	#      燃料温度和冷却剂温度密度写入OpenMC的xml文件
	#####################################################################################################

	##########   燃料  改 温度 组件模式
	tree = ET.parse('materials.xml')
	root = tree.getroot()

	# 遍历所有material元素
	materials = root.findall('material')
	#################################################################################################################
	#  处理燃料温度
	#################################################################################################################
	# 只处理fuel_1材料
	fuel_1_materials = [m for m in materials if m.get('name') == fuel_1_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_1_materials):  # 确保索引不越界
				fuel_1_materials[i].set('temperature', str(fuel_1_assembly_T[k]))
		j += n_pin_in_assembly
		k += 1

	# 只处理fuel_2材料
	fuel_2_materials = [m for m in materials if m.get('name') == fuel_2_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_2_materials):  # 确保索引不越界
				fuel_2_materials[i].set('temperature', str(fuel_2_assembly_T[k]))
		j += n_pin_in_assembly
		k += 1

	# 只处理fuel_3材料
	fuel_3_materials = [m for m in materials if m.get('name') == fuel_3_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_3_materials):  # 确保索引不越界
				fuel_3_materials[i].set('temperature', str(fuel_3_assembly_T[k]))
		j += n_pin_in_assembly
		k += 1

	# 只处理fuel_4材料
	fuel_4_materials = [m for m in materials if m.get('name') == fuel_4_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_4_materials):  # 确保索引不越界
				fuel_4_materials[i].set('temperature', str(fuel_4_assembly_T[k]))
		j += n_pin_in_assembly
		k += 1

	# 只处理fuel_5材料
	fuel_5_materials = [m for m in materials if m.get('name') == fuel_5_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_5_materials):  # 确保索引不越界
				fuel_5_materials[i].set('temperature', str(fuel_5_assembly_T[k]))
		j += n_pin_in_assembly
		k += 1

	#################################################################################################################
	#  处理燃料组件棒栅元内的冷却剂温度
	#################################################################################################################
	# 只处理fuel_1材料
	fuel_1_assembly_coolant_i_materials = [m for m in materials if m.get('name') == fuel_1_assembly_inner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_1_assembly_coolant_i_materials):  # 确保索引不越界
				fuel_1_assembly_coolant_i_materials[i].set('temperature', str(fuel_1_assembly_coolant_T[k]))
				density = fuel_1_assembly_coolant_i_materials[i].find('density')
				density.set('value', str(fuel_1_assembly_coolant_D[k]))
		j += n_pin_in_assembly
		k += 1
	# 只处理fuel_2材料
	fuel_2_assembly_coolant_i_materials = [m for m in materials if m.get('name') == fuel_2_assembly_inner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_2_assembly_coolant_i_materials):  # 确保索引不越界
				fuel_2_assembly_coolant_i_materials[i].set('temperature', str(fuel_2_assembly_coolant_T[k]))
				density = fuel_2_assembly_coolant_i_materials[i].find('density')
				density.set('value', str(fuel_2_assembly_coolant_D[k]))
		j += n_pin_in_assembly
		k += 1
	# 只处理fuel_3材料
	fuel_3_assembly_coolant_i_materials = [m for m in materials if m.get('name') == fuel_3_assembly_inner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_3_assembly_coolant_i_materials):  # 确保索引不越界
				fuel_3_assembly_coolant_i_materials[i].set('temperature', str(fuel_3_assembly_coolant_T[k]))
				density = fuel_3_assembly_coolant_i_materials[i].find('density')
				density.set('value', str(fuel_3_assembly_coolant_D[k]))
		j += n_pin_in_assembly
		k += 1

	# 只处理fuel_4材料
	fuel_4_assembly_coolant_i_materials = [m for m in materials if m.get('name') == fuel_4_assembly_inner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_4_assembly_coolant_i_materials):  # 确保索引不越界
				fuel_4_assembly_coolant_i_materials[i].set('temperature', str(fuel_4_assembly_coolant_T[k]))
				density = fuel_4_assembly_coolant_i_materials[i].find('density')
				density.set('value', str(fuel_4_assembly_coolant_D[k]))
		j += n_pin_in_assembly
		k += 1

	# 只处理fuel_5材料
	fuel_5_assembly_coolant_i_materials = [m for m in materials if m.get('name') == fuel_5_assembly_inner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_pin_in_assembly * n_fuel_assembly * n_floor :
		for i in range(j, j+n_pin_in_assembly):
			if i < len(fuel_5_assembly_coolant_i_materials):  # 确保索引不越界
				fuel_5_assembly_coolant_i_materials[i].set('temperature', str(fuel_5_assembly_coolant_T[k]))
				density = fuel_5_assembly_coolant_i_materials[i].find('density')
				density.set('value', str(fuel_5_assembly_coolant_D[k]))
		j += n_pin_in_assembly
		k += 1
	#################################################################################################################
	#  处理燃料组件边通道内的冷却剂温度和密度
	#################################################################################################################
	# 只处理fuel_1材料
	fuel_1_assembly_coolant_e_materials = [m for m in materials if m.get('name') == fuel_1_assembly_edge_coolant_channals_name]
	j = 0
	k = 0
	while j < n_edge_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_edge_channels_in_assembly):
			if i < len(fuel_1_assembly_coolant_e_materials):  # 确保索引不越界
				fuel_1_assembly_coolant_e_materials[i].set('temperature', str(fuel_1_assembly_coolant_T[k]))
				density = fuel_1_assembly_coolant_e_materials[i].find('density')
				density.set('value', str(fuel_1_assembly_coolant_D[k]))
		j += n_edge_channels_in_assembly
		k += 1
	# 只处理fuel_2材料
	fuel_2_assembly_coolant_e_materials = [m for m in materials if m.get('name') == fuel_2_assembly_edge_coolant_channals_name]
	j = 0
	k = 0
	while j < n_edge_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_edge_channels_in_assembly):
			if i < len(fuel_2_assembly_coolant_e_materials):  # 确保索引不越界
				fuel_2_assembly_coolant_e_materials[i].set('temperature', str(fuel_2_assembly_coolant_T[k]))
				density = fuel_2_assembly_coolant_e_materials[i].find('density')
				density.set('value', str(fuel_2_assembly_coolant_D[k]))
		j += n_edge_channels_in_assembly
		k += 1
	# 只处理fuel_3材料
	fuel_3_assembly_coolant_e_materials = [m for m in materials if m.get('name') == fuel_3_assembly_edge_coolant_channals_name]
	j = 0
	k = 0
	while j < n_edge_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_edge_channels_in_assembly):
			if i < len(fuel_3_assembly_coolant_e_materials):  # 确保索引不越界
				fuel_3_assembly_coolant_e_materials[i].set('temperature', str(fuel_3_assembly_coolant_T[k]))
				density = fuel_3_assembly_coolant_e_materials[i].find('density')
				density.set('value', str(fuel_3_assembly_coolant_D[k]))
		j += n_edge_channels_in_assembly
		k += 1

	# 只处理fuel_4材料
	fuel_4_assembly_coolant_e_materials = [m for m in materials if m.get('name') == fuel_4_assembly_edge_coolant_channals_name]
	j = 0
	k = 0
	while j < n_edge_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_edge_channels_in_assembly):
			if i < len(fuel_4_assembly_coolant_e_materials):  # 确保索引不越界
				fuel_4_assembly_coolant_e_materials[i].set('temperature', str(fuel_4_assembly_coolant_T[k]))
				density = fuel_4_assembly_coolant_e_materials[i].find('density')
				density.set('value', str(fuel_4_assembly_coolant_D[k]))
		j += n_edge_channels_in_assembly
		k += 1

	# 只处理fuel_5材料
	fuel_5_assembly_coolant_e_materials = [m for m in materials if m.get('name') == fuel_5_assembly_edge_coolant_channals_name]
	j = 0
	k = 0
	while j < n_edge_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_edge_channels_in_assembly):
			if i < len(fuel_5_assembly_coolant_e_materials):  # 确保索引不越界
				fuel_5_assembly_coolant_e_materials[i].set('temperature', str(fuel_5_assembly_coolant_T[k]))
				density = fuel_5_assembly_coolant_e_materials[i].find('density')
				density.set('value', str(fuel_5_assembly_coolant_D[k]))
		j += n_edge_channels_in_assembly
		k += 1
	#################################################################################################################
	#  处理燃料组件角通道内的冷却剂温度
	#################################################################################################################
	# 只处理fuel_1材料
	fuel_1_assembly_coolant_c_materials = [m for m in materials if m.get('name') == fuel_1_assembly_corner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_corner_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_corner_channels_in_assembly):
			if i < len(fuel_1_assembly_coolant_c_materials):  # 确保索引不越界
				fuel_1_assembly_coolant_c_materials[i].set('temperature', str(fuel_1_assembly_coolant_T[k]))
				density = fuel_1_assembly_coolant_c_materials[i].find('density')
				density.set('value', str(fuel_1_assembly_coolant_D[k]))
		j += n_corner_channels_in_assembly
		k += 1
	# 只处理fuel_2材料
	fuel_2_assembly_coolant_c_materials = [m for m in materials if m.get('name') == fuel_2_assembly_corner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_corner_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_corner_channels_in_assembly):
			if i < len(fuel_2_assembly_coolant_c_materials):  # 确保索引不越界
				fuel_2_assembly_coolant_c_materials[i].set('temperature', str(fuel_2_assembly_coolant_T[k]))
				density = fuel_2_assembly_coolant_c_materials[i].find('density')
				density.set('value', str(fuel_2_assembly_coolant_D[k]))
		j += n_corner_channels_in_assembly
		k += 1
	# 只处理fuel_3材料
	fuel_3_assembly_coolant_c_materials = [m for m in materials if m.get('name') == fuel_3_assembly_corner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_corner_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_corner_channels_in_assembly):
			if i < len(fuel_3_assembly_coolant_c_materials):  # 确保索引不越界
				fuel_3_assembly_coolant_c_materials[i].set('temperature', str(fuel_3_assembly_coolant_T[k]))
				density = fuel_3_assembly_coolant_c_materials[i].find('density')
				density.set('value', str(fuel_3_assembly_coolant_D[k]))
		j += n_corner_channels_in_assembly
		k += 1

	# 只处理fuel_4材料
	fuel_4_assembly_coolant_c_materials = [m for m in materials if m.get('name') == fuel_4_assembly_corner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_corner_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_corner_channels_in_assembly):
			if i < len(fuel_4_assembly_coolant_c_materials):  # 确保索引不越界
				fuel_4_assembly_coolant_c_materials[i].set('temperature', str(fuel_4_assembly_coolant_T[k]))
				density = fuel_4_assembly_coolant_c_materials[i].find('density')
				density.set('value', str(fuel_4_assembly_coolant_D[k]))
		j += n_corner_channels_in_assembly
		k += 1

	# 只处理fuel_5材料
	fuel_5_assembly_coolant_c_materials = [m for m in materials if m.get('name') == fuel_5_assembly_corner_coolant_channals_name]
	j = 0
	k = 0
	while j < n_corner_channels_in_assembly * n_fuel_assembly * n_floor:
		for i in range(j, j + n_corner_channels_in_assembly):
			if i < len(fuel_5_assembly_coolant_c_materials):  # 确保索引不越界
				fuel_5_assembly_coolant_c_materials[i].set('temperature', str(fuel_5_assembly_coolant_T[k]))
				density = fuel_5_assembly_coolant_c_materials[i].find('density')
				density.set('value', str(fuel_5_assembly_coolant_D[k]))
		j += n_corner_channels_in_assembly
		k += 1

	#写入到新文件
	tree.write('materials.xml', encoding='utf-8')

	#################################################################################################################

	print('物理程序运行...')
	os.system(run_openmc)

	##################################################################################################
	# 热工和物理数据保存
	##################################################################################################
	print("迭代数据转存...")
	directory = "./" + "迭代步"+str(iterations)+"输出"
	if not os.path.exists(directory):  # 文件夹存在检测并创建
		os.makedirs(directory)
	shutil.copy2("./materials.xml", "./" + "迭代步"+str(iterations)+"输出" + "/materials.xml")
	shutil.copy2("./geometry.xml", "./" + "迭代步"+str(iterations)+"输出" + "/geometry.xml")
	shutil.copy2("./settings.xml", "./" + "迭代步"+str(iterations)+"输出" + "/settings.xml")
	shutil.copy2("./tallies.xml", "./" + "迭代步"+str(iterations)+"输出" + "/tallies.xml")
	shutil.copy2(statepoint_file, "./" + "迭代步"+str(iterations)+"输出" + "/"+statepoint_file)
	shutil.copy2(cobra_input_file, "./" + "迭代步"+str(iterations)+"输出" +"/"+cobra_input_file) # 存储上一步的热工输入，为当前计算openmc使用的参数
	shutil.copy2(fuel_file, "./" + "迭代步"+str(iterations)+"输出" + "/"+fuel_file)
	shutil.copy2(coolant_file, "./" + "迭代步"+str(iterations)+"输出" + "/"+coolant_file)
	shutil.copy2(cobra_out_file, "./" + "迭代步"+str(iterations)+"输出" + "/"+cobra_out_file)
	np.savetxt("./" + "迭代步"+str(iterations)+"输出" +"/core_fuel_temp.csv", all_fuel_T_xH_yN, delimiter=', ')
	##################################################################################################################
	# 建立一维功率系数数组
	##################################################################################################################
	with h5py.File(statepoint_file, 'r') as hdf5_file:
		# data = np.zeros((max(n_fuel_diff_type) * n_eq_pin_in_assembly, n_floor * n_type_fuel))
		data_0 = np.array(hdf5_file['tallies/tally ' + str(fuel_1_start_tally_number) + '/results'][:, 0])[:, 0]
		data = data_0[data_0 != 0]  # 去除0元素

		# rows = data.shape
		# print(data)
		# print(f"数组有 {rows} (行，列)")
		##########################################################
		# 一维列功率系数数组归一化为一维功率数组
		##########################################################
		data_c = data  # rows行，columns为列
		data_c = np.array(data_c)

		data_c /= np.mean(data_c[data_c > 0.])  # 归一化，note：燃耗区不用相等，由于openmc计数卡计数的是不同区域的裂变次数，且仅需归一化出不同区域的功率
		data_c = core_power / (n_floor * n_fuel_assembly) * data_c  # 节块功率

	# data_c = np.array([f"{x:.5e}" for x in data_c])	#格式化数据
	# rows = data_c.shape
	# print(data_c)
	# print(f"数组有 {rows} 行，列")
	##################################################################################################
	# 组件线功率分层写入cobra输入卡
	##################################################################################################
	floor = 1
	i = 0
	assembly_powers = np.zeros((int(len(data_c) / n_floor), n_floor))
	assembly_liner_powers = np.zeros((int(len(data_c) / n_floor), n_floor))
	while floor <= n_floor:
		assembly_powers[:, floor - 1] = data_c[int(i):int(i + len(data_c) / n_floor)]
		assembly_liner_powers[:, floor - 1] = assembly_powers[:, floor - 1] / h_pin[floor - 1]
		#assembly_liner_powers[:, floor - 1][half_assembly_indices] *= 2
		#assembly_liner_powers[:, floor - 1][quarter_assembly_indices] *= 4
		i += len(data_c) / n_floor
		floor += 1

	assembly_liner_powers = np.array(
		[[f"{x:.5e}" for x in row] for row in assembly_liner_powers])  # 注意：此为二维数据格式化,科学计数法且保留五位小数

	# print(assembly_liner_powers[:,0])

	print('耦合模式：' + str(coupled_mode))

	with open(cobra_input_file, "r") as file:
		lines = file.readlines()

	########################################################################  assembly

	# 找到 "Start Line" 和 "End Line" 的位置
	start_index = None
	end_index = None
	for i, line in enumerate(lines):
		if line.strip() == liner_power_start_insert_point:
			start_index = i
		elif line.strip() == liner_power_end_insert_point:
			end_index = i
	if start_index is None or end_index is None:  # 检查是否找到两行
		print("未找到 'Start Line' 或 'End Line'")
	else:
		floor = 1
		new_content = ""  # 构造新的内容
		new_content = new_content + "    "+str(n_floor)+ "\n"
		while floor <= n_floor:
			new_content = new_content + "  " + str(z_pin[floor-1]) + "\n"
			for i, item in enumerate(assembly_liner_powers[:, floor - 1]):
				new_content += str(item) + " "
				if (i + 1) % 6 == 0:  # 每六个元素一换行
					new_content += "\n"
			if n_fuel_assembly % 6 != 0:
				new_content += "\n"
			floor += 1
		lines[start_index + 1:end_index] = new_content
	
	with open(cobra_input_file, "w") as file:  # 写入输出文件
		file.writelines(lines)
		print(f"功率信息载入...")
	
	#####################################################################################################
	#              运行cobra
	#####################################################################################################
	os.system(run_cobra)
	#####################################################################################################
	#              读取燃料温度建立数组
	#####################################################################################################
	# 读取文件内容
	with open(fuel_file, 'r') as file:
		file_content = file.read()
	
	# 初始化变量
	rods_data = {}  # 存储每个燃料棒的温度数据
	current_rod = None
	height_positions = []  # 存储所有高度位置
	
	# 正则表达式模式
	rod_pattern = re.compile(r'TEMPERATURE DATA FOR ROD\s+(\d+) \(FUEL TYPE\s+\d+\)')
	temp_pattern = re.compile(r'\s*(\d+\.\d+)\s+\d+\.\d+\s+\d+\.\d+\s+\d+\s+(\d+\.\d+)')
	
	# 处理每一行
	for line in file_content.split('\n'):
	# 检查是否是新的燃料棒数据开始
		rod_match = rod_pattern.search(line)
		if rod_match:
			current_rod = int(rod_match.group(1))
			rods_data[current_rod] = {}  # 初始化该燃料棒的数据字典
			continue
	
	# 检查是否是温度数据行
		temp_match = temp_pattern.match(line)
		if temp_match and current_rod is not None:
			height = temp_match.group(1)
			temperature = float(temp_match.group(2))
	
			# 记录高度位置（如果尚未记录）
			if height not in height_positions:
				height_positions.append(height)
	
			rods_data[current_rod][height] = temperature  # 存储该燃料棒在该高度的温度
	
	# 按高度排序（从小到大）
	height_positions = sorted(height_positions, key=float)
	
	# 按燃料棒编号排序（从小到大）
	rod_numbers = sorted(rods_data.keys())
	
	# 构建二维数组（行=燃料棒，列=高度）
	temperature_2d_array = []
	for rod in rod_numbers:
		rod_temps = []
		for height in height_positions:
			rod_temps.append(rods_data[rod][height])
		temperature_2d_array.append(rod_temps)
	
	#温度单位由摄氏度转为K
	fuel_temperature_2d_array = np.array(temperature_2d_array)
	fuel_temperature_2d_array = np.array([temp + temp_c for temp in fuel_temperature_2d_array])

	#####################################################################################################
	#      读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理
	#####################################################################################################
	floor = 1
	fuel_1_assembly_T = []  # rows行，columns为列
	fuel_2_assembly_T = []  # rows行，columns为列
	fuel_3_assembly_T = []  # rows行，columns为列
	fuel_4_assembly_T = []  # rows行，columns为列
	fuel_5_assembly_T = []  # rows行，columns为列
	while floor <= n_floor:
		i = 0
		j = 0
		while i <= len(core_map) - 1:
			if core_map[i] == 0:
				print("have no fuel region")
			elif core_map[i] == 1:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = fuel_temperature_2d_array[j, floor - 1]
					fuel_1_assembly_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 2:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = fuel_temperature_2d_array[j, floor - 1]
					fuel_2_assembly_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 3:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = fuel_temperature_2d_array[j, floor - 1]
					fuel_3_assembly_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 4:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = fuel_temperature_2d_array[j, floor - 1]
					fuel_4_assembly_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 5:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = fuel_temperature_2d_array[j, floor - 1]
					fuel_5_assembly_T.append(subset)
					k += 1
					j += 1
			else:
				print(f"读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理错误")
				sys.exit(1)
			i += 1
		floor += 1
	fuel_1_assembly_T = np.array(fuel_1_assembly_T)
	fuel_2_assembly_T = np.array(fuel_2_assembly_T)
	fuel_3_assembly_T = np.array(fuel_3_assembly_T)
	fuel_4_assembly_T = np.array(fuel_4_assembly_T)
	fuel_5_assembly_T = np.array(fuel_5_assembly_T)
	# print(fuel_1_assembly_T)
	# rows = fuel_1_assembly_T.shape
	# print(fuel_1_assembly_T)
	# print(f"数组有 {rows} 行，列")
	#####################################################################################################
	#              读取冷却剂温度和密度建立数组
	#####################################################################################################
	# 读取文件内容
	with open(coolant_file, 'r') as file:
		content = file.read()

	# 初始化数据结构
	channels = {}

	# 使用正则表达式分割不同通道的数据
	channel_sections = re.split(r'TIME =\s+\d+\.\d+ SEC\s+-\s+RESULTS FOR CHANNEL\s+\d+', content)[1:]

	# 遍历每个通道的数据
	for i, section in enumerate(channel_sections, start=1):
		# 初始化当前通道的数据列表
		channels[i] = {'distance': [], 'temperature': [], 'density': []}

		# 分割行并跳过标题行
		lines = [line.strip() for line in section.split('\n') if line.strip()]
		data_lines = lines[2:]  # 跳过前两行（标题和空行）

		# 提取每行的数据
		for line in data_lines:
			# 使用正则表达式匹配数据行
			match = re.match(r'\s*(\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)\s+(-?\d+\.\d+)', line)
			if match:
				distance = float(match.group(1))
				temperature = float(match.group(4))
				density = float(match.group(5))

				channels[i]['distance'].append(distance)
				channels[i]['temperature'].append(temperature)
				channels[i]['density'].append(density)

	# 创建二维数组（假设所有通道有相同数量的数据点）
	num_channels = len(channels)
	num_points = len(channels[1]['temperature'])

	# 温度二维数组 (channels × points)
	coolant_temperature_2d = []
	for ch in range(1, num_channels + 1):
		coolant_temperature_2d.append([channels[ch]['temperature'][i] for i in range(num_points)])

	# 密度二维数组 (channels × points)
	coolant_density_2d = []
	for ch in range(1, num_channels + 1):
		coolant_density_2d.append([channels[ch]['density'][i] for i in range(num_points)])

	# 转nump数组温度单位由摄氏度转为K
	coolant_temperature_2d_array = np.array(coolant_temperature_2d)
	coolant_temperature_2d_array = np.array([temp + temp_c for temp in coolant_temperature_2d_array])
	# 转nump数组
	coolant_density_2d_array = np.array(coolant_density_2d)

	# 平均值法求解节块中心温度
	coolant_temperature_2d_array = (coolant_temperature_2d_array[:, :-1] + coolant_temperature_2d_array[:, 1:]) / 2
	# 平均值法求解节块中心密度
	coolant_density_2d_array = (coolant_density_2d_array[:, :-1] + coolant_density_2d_array[:, 1:]) / 2
	# 密度单位kg/m3改为g/cm3
	coolant_density_2d_array = coolant_density_2d_array / 1000

	#####################################################################################################
	#      读取冷却剂温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理
	#####################################################################################################
	floor = 1
	fuel_1_assembly_coolant_T = []  # rows行，columns为列
	fuel_2_assembly_coolant_T = []  # rows行，columns为列
	fuel_3_assembly_coolant_T = []  # rows行，columns为列
	fuel_4_assembly_coolant_T = []  # rows行，columns为列
	fuel_5_assembly_coolant_T = []  # rows行，columns为列
	while floor <= n_floor:
		i = 0
		j = 0
		while i <= len(core_map) - 1:
			if core_map[i] == 0:
				print("have no fuel region")
			elif core_map[i] == 1:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_temperature_2d_array[j, floor - 1]
					fuel_1_assembly_coolant_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 2:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_temperature_2d_array[j, floor - 1]
					fuel_2_assembly_coolant_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 3:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_temperature_2d_array[j, floor - 1]
					fuel_3_assembly_coolant_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 4:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_temperature_2d_array[j, floor - 1]
					fuel_4_assembly_coolant_T.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 5:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_temperature_2d_array[j, floor - 1]
					fuel_5_assembly_coolant_T.append(subset)
					k += 1
					j += 1
			else:
				print(f"读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理错误")
				sys.exit(1)
			i += 1
		floor += 1
	fuel_1_assembly_coolant_T = np.array(fuel_1_assembly_coolant_T)
	fuel_2_assembly_coolant_T = np.array(fuel_2_assembly_coolant_T)
	fuel_3_assembly_coolant_T = np.array(fuel_3_assembly_coolant_T)
	fuel_4_assembly_coolant_T = np.array(fuel_4_assembly_coolant_T)
	fuel_5_assembly_coolant_T = np.array(fuel_5_assembly_coolant_T)
	#####################################################################################################
	#      读取冷却剂密度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理
	#####################################################################################################
	floor = 1
	fuel_1_assembly_coolant_D = []  # rows行，columns为列
	fuel_2_assembly_coolant_D = []  # rows行，columns为列
	fuel_3_assembly_coolant_D = []  # rows行，columns为列
	fuel_4_assembly_coolant_D = []  # rows行，columns为列
	fuel_5_assembly_coolant_D = []  # rows行，columns为列
	while floor <= n_floor:
		i = 0
		j = 0
		while i <= len(core_map) - 1:
			if core_map[i] == 0:
				print("have no fuel region")
			elif core_map[i] == 1:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_density_2d_array[j, floor - 1]
					fuel_1_assembly_coolant_D.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 2:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_density_2d_array[j, floor - 1]
					fuel_2_assembly_coolant_D.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 3:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_density_2d_array[j, floor - 1]
					fuel_3_assembly_coolant_D.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 4:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_density_2d_array[j, floor - 1]
					fuel_4_assembly_coolant_D.append(subset)
					k += 1
					j += 1
			elif core_map[i] == 5:
				k = 0
				while k < n_eq_pin_in_assembly:
					subset = coolant_density_2d_array[j, floor - 1]
					fuel_5_assembly_coolant_D.append(subset)
					k += 1
					j += 1
			else:
				print(f"读取燃料温度建立的数组按照堆芯燃料类型和排序（coremap）规范化处理错误")
				sys.exit(1)
			i += 1
		floor += 1
	fuel_1_assembly_coolant_D = np.array(fuel_1_assembly_coolant_D)
	fuel_2_assembly_coolant_D = np.array(fuel_2_assembly_coolant_D)
	fuel_3_assembly_coolant_D = np.array(fuel_3_assembly_coolant_D)
	fuel_4_assembly_coolant_D = np.array(fuel_4_assembly_coolant_D)
	fuel_5_assembly_coolant_D = np.array(fuel_5_assembly_coolant_D)

	#####################################################################################################
	#      燃料和冷却剂信息转存
	#####################################################################################################
	fuel_1_assembly_T1 = fuel_1_assembly_T
	fuel_2_assembly_T1 = fuel_2_assembly_T
	fuel_3_assembly_T1 = fuel_3_assembly_T
	fuel_4_assembly_T1 = fuel_4_assembly_T
	fuel_5_assembly_T1 = fuel_5_assembly_T

	fuel_1_assembly_coolant_T1 = fuel_1_assembly_coolant_T
	fuel_2_assembly_coolant_T1 = fuel_2_assembly_coolant_T
	fuel_3_assembly_coolant_T1 = fuel_3_assembly_coolant_T
	fuel_4_assembly_coolant_T1 = fuel_4_assembly_coolant_T
	fuel_5_assembly_coolant_T1 = fuel_5_assembly_coolant_T

	fuel_1_assembly_coolant_D1 = fuel_1_assembly_coolant_D
	fuel_2_assembly_coolant_D1 = fuel_2_assembly_coolant_D
	fuel_3_assembly_coolant_D1 = fuel_3_assembly_coolant_D
	fuel_4_assembly_coolant_D1 = fuel_4_assembly_coolant_D
	fuel_5_assembly_coolant_D1 = fuel_5_assembly_coolant_D
	# print(fuel_1_assembly_T)
	#####################################################################################################
	#      燃料和冷却剂的picard迭代
	#####################################################################################################
	fuel_1_assembly_T = fuel_1_assembly_T0 + relaxation_factor * (fuel_1_assembly_T1 - fuel_1_assembly_T0)
	fuel_2_assembly_T = fuel_2_assembly_T0 + relaxation_factor * (fuel_2_assembly_T1 - fuel_2_assembly_T0)
	fuel_3_assembly_T = fuel_3_assembly_T0 + relaxation_factor * (fuel_3_assembly_T1 - fuel_3_assembly_T0)
	fuel_4_assembly_T = fuel_4_assembly_T0 + relaxation_factor * (fuel_4_assembly_T1 - fuel_4_assembly_T0)
	fuel_5_assembly_T = fuel_5_assembly_T0 + relaxation_factor * (fuel_5_assembly_T1 - fuel_5_assembly_T0)

	fuel_1_assembly_coolant_T = fuel_1_assembly_coolant_T0 + relaxation_factor * (fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0)
	fuel_2_assembly_coolant_T = fuel_2_assembly_coolant_T0 + relaxation_factor * (fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0)
	fuel_3_assembly_coolant_T = fuel_3_assembly_coolant_T0 + relaxation_factor * (fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0)
	fuel_4_assembly_coolant_T = fuel_4_assembly_coolant_T0 + relaxation_factor * (fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0)
	fuel_5_assembly_coolant_T = fuel_5_assembly_coolant_T0 + relaxation_factor * (fuel_5_assembly_coolant_T1 - fuel_5_assembly_coolant_T0)

	fuel_1_assembly_coolant_D = fuel_1_assembly_coolant_D0 + relaxation_factor * (fuel_1_assembly_coolant_D1 - fuel_1_assembly_coolant_D0)
	fuel_2_assembly_coolant_D = fuel_2_assembly_coolant_D0 + relaxation_factor * (fuel_2_assembly_coolant_D1 - fuel_2_assembly_coolant_D0)
	fuel_3_assembly_coolant_D = fuel_3_assembly_coolant_D0 + relaxation_factor * (fuel_3_assembly_coolant_D1 - fuel_3_assembly_coolant_D0)
	fuel_4_assembly_coolant_D = fuel_4_assembly_coolant_D0 + relaxation_factor * (fuel_4_assembly_coolant_D1 - fuel_4_assembly_coolant_D0)
	fuel_5_assembly_coolant_D = fuel_5_assembly_coolant_D0 + relaxation_factor * (fuel_5_assembly_coolant_D1 - fuel_5_assembly_coolant_D0)


	print('======================第'+str(iterations)+'次迭代结束==========================')
	#print(fuel_1_assembly_T)
	# 定义各个条件
	if n_type_fuel == 1:
		fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_temp_condition_max = fuel_condition1
		coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_temp_condition_max = coolant_condition1

		fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1,))
		fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0,))
		fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
		fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
		coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1,))
		coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0,))
		coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
		coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
		max_fuel_convergence = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf)
		max_coolant_convergence = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf)
	elif n_type_fuel == 2:
		fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_temp_condition_max = fuel_condition1 or fuel_condition2
		coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_temp_condition_max = coolant_condition1 or coolant_condition2

		fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1, fuel_2_assembly_T1))
		fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0, fuel_2_assembly_T0))
		fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
		fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
		coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1))
		coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0))
		coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
		coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
		max_fuel_convergence = max(
			np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf))
		max_coolant_convergence = max(
			np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf))
	elif n_type_fuel == 3:
		fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition3 = np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_temp_condition_max = fuel_condition1 or fuel_condition2 or fuel_condition3
		coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition3 = np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_temp_condition_max = coolant_condition1 or coolant_condition2 or coolant_condition3

		fuel_all_assembly_T1 = np.concatenate((fuel_1_assembly_T1, fuel_2_assembly_T1, fuel_3_assembly_T1))
		fuel_all_assembly_T0 = np.concatenate((fuel_1_assembly_T0, fuel_2_assembly_T0, fuel_3_assembly_T0))
		fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
		fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
		coolant_all_assembly_T1 = np.concatenate(
			(fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1, fuel_3_assembly_coolant_T1))
		coolant_all_assembly_T0 = np.concatenate(
			(fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0, fuel_3_assembly_coolant_T0))
		coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
		coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
		max_fuel_convergence = max(
			np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf))
		max_coolant_convergence = max(
			np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0, ord=np.inf))
	elif n_type_fuel == 4:
		fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition3 = np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition4 = np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_temp_condition_max = fuel_condition1 or fuel_condition2 or fuel_condition3 or fuel_condition4
		coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition3 = np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition4 = np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_temp_condition_max = coolant_condition1 or coolant_condition2 or coolant_condition3 or coolant_condition4

		fuel_all_assembly_T1 = np.concatenate(
			(fuel_1_assembly_T1, fuel_2_assembly_T1, fuel_3_assembly_T1, fuel_4_assembly_T1))
		fuel_all_assembly_T0 = np.concatenate(
			(fuel_1_assembly_T0, fuel_2_assembly_T0, fuel_3_assembly_T0, fuel_4_assembly_T0))
		fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
		fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
		coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1,
												  fuel_3_assembly_coolant_T1, fuel_4_assembly_coolant_T1))
		coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0,
												  fuel_3_assembly_coolant_T0, fuel_4_assembly_coolant_T0))
		coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
		coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
		max_fuel_convergence = max(
			np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0, ord=np.inf))
		max_coolant_convergence = max(
			np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0, ord=np.inf))
	elif n_type_fuel == 5:
		fuel_condition1 = np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition2 = np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition3 = np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition4 = np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_condition5 = np.linalg.norm(fuel_5_assembly_T1 - fuel_5_assembly_T0,
										 ord=np.inf) > fuel_temp_convergence_limit
		fuel_temp_condition_max = fuel_condition1 or fuel_condition2 or fuel_condition3 or fuel_condition4 or fuel_condition5
		coolant_condition1 = np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition2 = np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition3 = np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition4 = np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_condition5 = np.linalg.norm(fuel_5_assembly_coolant_T1 - fuel_5_assembly_coolant_T0,
											ord=np.inf) > coolant_temp_convergence_limit
		coolant_temp_condition_max = coolant_condition1 or coolant_condition2 or coolant_condition3 or coolant_condition4 or coolant_condition5

		fuel_all_assembly_T1 = np.concatenate(
			(fuel_1_assembly_T1, fuel_2_assembly_T1, fuel_3_assembly_T1, fuel_4_assembly_T1, fuel_5_assembly_T1))
		fuel_all_assembly_T0 = np.concatenate(
			(fuel_1_assembly_T0, fuel_2_assembly_T0, fuel_3_assembly_T0, fuel_4_assembly_T0, fuel_5_assembly_T0))
		fuel_temp_rmse = np.sqrt(np.mean(np.abs(fuel_all_assembly_T1 - fuel_all_assembly_T0) ** 2))
		fuel_temp_condition = fuel_temp_rmse > fuel_temp_convergence_limit
		coolant_all_assembly_T1 = np.concatenate((fuel_1_assembly_coolant_T1, fuel_2_assembly_coolant_T1,
												  fuel_3_assembly_coolant_T1, fuel_4_assembly_coolant_T1,
												  fuel_5_assembly_coolant_T1))
		coolant_all_assembly_T0 = np.concatenate((fuel_1_assembly_coolant_T0, fuel_2_assembly_coolant_T0,
												  fuel_3_assembly_coolant_T0, fuel_4_assembly_coolant_T0,
												  fuel_5_assembly_coolant_T0))
		coolant_temp_rmse = np.sqrt(np.mean(np.abs(coolant_all_assembly_T1 - coolant_all_assembly_T0) ** 2))
		coolant_temp_condition = coolant_temp_rmse > coolant_temp_convergence_limit
		max_fuel_convergence = max(
			np.linalg.norm(fuel_1_assembly_T1 - fuel_1_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_T1 - fuel_2_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_3_assembly_T1 - fuel_3_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_4_assembly_T1 - fuel_4_assembly_T0, ord=np.inf),
			np.linalg.norm(fuel_5_assembly_T1 - fuel_5_assembly_T0, ord=np.inf))
		max_coolant_convergence = max(
			np.linalg.norm(fuel_1_assembly_coolant_T1 - fuel_1_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_2_assembly_coolant_T1 - fuel_2_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_3_assembly_coolant_T1 - fuel_3_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_4_assembly_coolant_T1 - fuel_4_assembly_coolant_T0, ord=np.inf),
			np.linalg.norm(fuel_5_assembly_coolant_T1 - fuel_5_assembly_coolant_T0, ord=np.inf))
	else:
		print("收敛条件处理错误")

	abs_error = np.abs(assembly_powers - assembly_powers0)
	eps = 1e-12
	relative_error = abs_error / (np.abs(assembly_powers0) + eps)  # 计算相对误差（避免除以零）


	abs_error = np.abs(assembly_powers - assembly_powers0)
	eps = 1e-12
	relative_error = abs_error / (np.abs(assembly_powers0) + eps)  # 计算相对误差（避免除以零）
	power_rmse = np.sqrt(np.mean(relative_error ** 2))

	power_condition_max = np.max(relative_error) > power_convergence_limit
	power_condition = power_rmse > power_convergence_limit

	print("最大范数数据处理：")
	print(" 燃料温度收敛判据:", max_fuel_convergence)
	print(" 冷却剂温度收敛判据:", max_coolant_convergence)
	print(" 功率收敛判据:", np.max(relative_error))

	print("均方根数据处理：")
	print(" 燃料温度收敛判据:", fuel_temp_rmse)
	print(" 冷却剂温度收敛判据:", coolant_temp_rmse)
	print(" 功率收敛判据:", power_rmse)

	all_convergence = np.array([[max_fuel_convergence,fuel_temp_rmse],
					   [max_coolant_convergence,coolant_temp_rmse],
					   [np.max(relative_error),power_rmse]])
	np.savetxt("./" + "迭代步" + str(iterations) + "输出" + "/收敛判据.csv", all_convergence, delimiter=', ')
	np.savetxt("./" + "迭代步" + str(iterations) + "输出" + "/assembly_powers.csv", assembly_powers, delimiter=', ')
	
	wb = Workbook()
	ws = wb.active
	# 将三个列表分别写入Excel的三列
	for i in range(len(fuel_1_assembly_T1)):
		ws.cell(row=i+1, column=1, value=fuel_1_assembly_T1[i])
	for i in range(len(fuel_2_assembly_T1)):
		ws.cell(row=i+1, column=2, value=fuel_2_assembly_T1[i])
	for i in range(len(fuel_3_assembly_T1)):
		ws.cell(row=i+1, column=3, value=fuel_3_assembly_T1[i])
	for i in range(len(fuel_4_assembly_T1)):
		ws.cell(row=i+1, column=4, value=fuel_4_assembly_T1[i])
	for i in range(len(fuel_5_assembly_T1)):
		ws.cell(row=i+1, column=5, value=fuel_5_assembly_T1[i])
	for i in range(len(fuel_1_assembly_coolant_T1)):
		ws.cell(row=i+1, column=6, value=fuel_1_assembly_coolant_T1[i])
	for i in range(len(fuel_2_assembly_coolant_T1)):
		ws.cell(row=i+1, column=7, value=fuel_2_assembly_coolant_T1[i])
	for i in range(len(fuel_3_assembly_coolant_T1)):
		ws.cell(row=i+1, column=8, value=fuel_3_assembly_coolant_T1[i])
	for i in range(len(fuel_4_assembly_coolant_T1)):
		ws.cell(row=i+1, column=9, value=fuel_4_assembly_coolant_T1[i])
	for i in range(len(fuel_5_assembly_coolant_T1)):
		ws.cell(row=i+1, column=10, value=fuel_5_assembly_coolant_T1[i])
	for i in range(len(fuel_1_assembly_coolant_D1)):
		ws.cell(row=i+1, column=11, value=fuel_1_assembly_coolant_D1[i])
	for i in range(len(fuel_2_assembly_coolant_D1)):
		ws.cell(row=i+1, column=12, value=fuel_2_assembly_coolant_D1[i])
	for i in range(len(fuel_3_assembly_coolant_D1)):
		ws.cell(row=i+1, column=13, value=fuel_3_assembly_coolant_D1[i])
	for i in range(len(fuel_4_assembly_coolant_D1)):
		ws.cell(row=i+1, column=14, value=fuel_4_assembly_coolant_D1[i])
	for i in range(len(fuel_5_assembly_coolant_D1)):
		ws.cell(row=i+1, column=15, value=fuel_5_assembly_coolant_D1[i])
	
	# 保存Excel文件
	wb.save("./" + "迭代步" + str(iterations) + "输出" + "/assembly_T_D.xlsx")

	iterations += 1

# 打印结果
print("\n========================最终结果============================")
# 创建一个新的工作簿
wb = Workbook()
ws = wb.active
# 将三个列表分别写入Excel的三列
for i in range(len(fuel_1_assembly_T0)):
	ws.cell(row=i+1, column=1, value=fuel_1_assembly_T0[i])
for i in range(len(fuel_2_assembly_T0)):
	ws.cell(row=i+1, column=2, value=fuel_2_assembly_T0[i])
for i in range(len(fuel_3_assembly_T0)):
	ws.cell(row=i+1, column=2, value=fuel_3_assembly_T0[i])
for i in range(len(fuel_4_assembly_T0)):
	ws.cell(row=i+1, column=1, value=fuel_4_assembly_T0[i])
for i in range(len(fuel_4_assembly_T0)):
	ws.cell(row=i+1, column=2, value=fuel_5_assembly_T0[i])
# 保存Excel文件
wb.save('FUEL_T.xlsx')

# print("高度位置:", height_positions)
# print("燃料组件编号:", rod_numbers)
# print("\n二维数组（行=燃料组件，列=高度，单位：K）:")
# for i, rod_temps in enumerate(all_fuel_T_xH_yN):
# 	print(f"燃料组件 {rod_numbers[i]}: {rod_temps}")

#np.savetxt('core_fuel_temp.csv', all_fuel_T_xH_yN, delimiter=', ')

print('\n  normal end of execution for OpenMC-0.14 and COBRA')
print('  check for warning in listing')
print('  before assuming your run was successful')
